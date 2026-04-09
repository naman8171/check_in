from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook

class ImportLineWizard(models.TransientModel):
    _name = 'import.line.wizard'
    _description = 'Import Lines Wizard'

    upload_file = fields.Binary(string="Upload File", required=True)
    file_name = fields.Char(string="File Name")

    def action_import(self):

        if not self.upload_file:
            raise UserError("Please upload a file.")

        active_id = self.env.context.get('active_id')
        import_type = self.env.context.get('import_type')

        if not active_id:
            raise UserError("No active dossier found.")

        dossier = self.env['technical.dossier'].browse(active_id)

        file_data = base64.b64decode(self.upload_file)
        file_stream = io.BytesIO(file_data)
        workbook = load_workbook(file_stream)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            if import_type == 'cogs':

                cost_component = row_data.get('Cost Component')
                amount = row_data.get('Amount')

                try:
                    amount = float(amount) if amount not in (None, '', False) else 0.0
                except ValueError:
                    amount = 0.0

                self.env['cogs.analysis'].create({
                    'dossier_id': dossier.id,
                    'cost_component': cost_component or '',
                    'amount': amount,
                })

            elif import_type == 'final_formulation':

                ingredient = row_data.get('Ingredient')

                function = row_data.get('Function')

                percentage = (row_data.get('Percentage_w_w'))

                regulatory_class = row_data.get('Regulatory_Class')

                supplier_name = row_data.get('Supplier')

                cost_per_kg = row_data.get('Cost_Per_kg')

                try:
                    percentage = float(percentage) if percentage not in (None, '', False) else 0.0
                except ValueError:
                    percentage = 0.0

                try:
                    cost_per_kg = float(cost_per_kg) if cost_per_kg not in (None, '', False) else 0.0
                except ValueError:
                    cost_per_kg = 0.0

                self.env['final.formulation'].create({
                    'dossier_id': dossier.id,
                    'ingredient': ingredient or '',
                    'function': function or '',
                    'percentage_w_w': percentage,
                    'regulatory_class': regulatory_class or '',
                    'supplier': supplier_name or '',
                    'cost_per_kg': cost_per_kg,
                })

        return {'type': 'ir.actions.act_window_close'}
